from openpyxl import load_workbook
### aba fonte
### aba cadastro_nequipto
### aba cadastro_equipto
arquivo = load_workbook(r"fonte_dados.xlsx")

def transferir_dados_aba_destino(aba_origem,aba_destino,linha_origem):
    linha_destino = aba_destino.max_row + 1
    ## for x in (1,13):
    celula_origem = aba_origem.cell(row=linha_origem,column=1)
    celula_destino = aba_destino.cell(row=linha_destino,column=1)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=2)
    celula_destino = aba_destino.cell(row=linha_destino,column=2)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=3)
    celula_destino = aba_destino.cell(row=linha_destino,column=3)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=4)
    celula_destino = aba_destino.cell(row=linha_destino,column=4)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=5)
    celula_destino = aba_destino.cell(row=linha_destino,column=5)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=6)
    celula_destino = aba_destino.cell(row=linha_destino,column=6)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=7)
    celula_destino = aba_destino.cell(row=linha_destino,column=7)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=8)
    celula_destino = aba_destino.cell(row=linha_destino,column=8)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=9)
    celula_destino = aba_destino.cell(row=linha_destino,column=9)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=10)
    celula_destino = aba_destino.cell(row=linha_destino,column=10)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=11)
    celula_destino = aba_destino.cell(row=linha_destino,column=11)
    celula_destino.value = celula_origem.value
    celula_origem = aba_origem.cell(row=linha_origem,column=12)
    celula_destino = aba_destino.cell(row=linha_destino,column=12)
    celula_destino.value = celula_origem.value

##print(arquivo.sheetnames)
aba_atual = arquivo["fonte"]
#print(aba_atual)
# Selecionar uma ba especifica
ult_linha = aba_atual.max_row
#print(ult_linha)

linha = 2
coluna = 1
for linha in range(linha,ult_linha): # ###############
    coluna = 1 ############## Tirar este comentario depois
    if (aba_atual.cell(row=linha,column=coluna).value) == 'N':
        ##print("*************Não Equipto !******************")
        aba_destino = arquivo["cadastro_nequipto"]
        linha_destino = aba_destino.max_row + 1
        transferir_dados_aba_destino(aba_atual,aba_destino,linha)
    else:
        ##print("*************Equipto !******************")
        aba_destino = arquivo["cadastro_equipto"]
        linha_destino = aba_destino.max_row + 1
        transferir_dados_aba_destino(aba_atual,aba_destino,linha)

arquivo.save("fonte_dados2.xlsx")
